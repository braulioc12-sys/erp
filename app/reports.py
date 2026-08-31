"""Generación de reportes en Excel (.xlsx) usando openpyxl."""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.helpers import pretty_label

# Paleta acorde a la app (ver app/static/css/style.css)
COLOR_PRIMARY = "1D4ED8"
COLOR_PRIMARY_SOFT = "E8EDFC"
COLOR_HEADER_TEXT = "FFFFFF"
COLOR_GRAY = "667085"
COLOR_TOTAL_FILL = "111827"

CURRENCY_FORMAT = '"S/" #,##0.00'
COLUMN_WIDTHS = [13, 16, 12, 12, 44, 16]  # Fecha, Tipo, Viaje, Unidad, Descripción, Monto
COLUMNS = ["Fecha", "Tipo", "Viaje", "Unidad", "Descripción", "Monto"]


def _thin_border(sides="bottom"):
    side = Side(style="thin", color="D0D5DD")
    kwargs = {s: side for s in ("left", "right", "top", "bottom") if s in sides or sides == "all"}
    return Border(**kwargs)


def build_expenses_workbook(expenses, company_name, filter_description, known_type_order=None):
    """Construye un workbook con el reporte de gastos agrupado por tipo,
    con subtotales y total general. `expenses` es una lista de filas
    (sqlite3.Row) con: expense_date, type, trip_code, vehicle_plate,
    description, amount. `known_type_order` (opcional) es la lista de tipos
    en el orden del catálogo — los tipos que aparezcan en los gastos pero no
    en esta lista (por ejemplo un concepto ya desactivado) se agregan al
    final, en orden alfabético."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos"

    last_col_letter = get_column_letter(len(COLUMNS))

    # --- Título y subtítulo ---
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = f"{company_name} — Reporte de Gastos"
    ws["A1"].font = Font(bold=True, size=14, color=COLOR_PRIMARY)

    ws.merge_cells(f"A2:{last_col_letter}2")
    generated = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws["A2"] = f"Generado el {generated}  ·  {filter_description}"
    ws["A2"].font = Font(italic=True, size=10, color=COLOR_GRAY)

    header_row = 4

    # --- Encabezados de columna ---
    for idx, title in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=idx, value=title)
        cell.font = Font(bold=True, color=COLOR_HEADER_TEXT)
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="right" if title == "Monto" else "left", vertical="center")
        cell.border = _thin_border("all")

    ws.freeze_panes = f"A{header_row + 1}"

    # --- Agrupar gastos por tipo. El orden de los grupos sigue el orden en
    # que aparecen en el catálogo (app/routes/catalogos.py); si un gasto usa
    # un tipo que ya no está en el catálogo (por ejemplo fue desactivado),
    # igual se muestra, al final. ---
    by_type = {}
    for e in expenses:
        by_type.setdefault(e["type"], []).append(e)

    known_type_order = known_type_order or []
    type_order = [t for t in known_type_order if t in by_type]
    extra_types = sorted(t for t in by_type if t not in known_type_order)
    type_order += extra_types

    row = header_row + 1
    grand_total = 0.0

    for type_key in type_order:
        rows = by_type.get(type_key) or []
        if not rows:
            continue
        rows = sorted(rows, key=lambda r: r["expense_date"])

        # Subencabezado de grupo
        ws.merge_cells(f"A{row}:{last_col_letter}{row}")
        group_cell = ws.cell(row=row, column=1, value=pretty_label(type_key))
        group_cell.font = Font(bold=True, color=COLOR_PRIMARY)
        group_cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY_SOFT)
        row += 1

        subtotal = 0.0
        for e in rows:
            ws.cell(row=row, column=1, value=e["expense_date"])
            ws.cell(row=row, column=2, value=pretty_label(e["type"]))
            ws.cell(row=row, column=3, value=e["trip_code"] or "—")
            ws.cell(row=row, column=4, value=e["vehicle_plate"] or "—")
            ws.cell(row=row, column=5, value=e["description"] or "")
            amount_cell = ws.cell(row=row, column=6, value=float(e["amount"]))
            amount_cell.number_format = CURRENCY_FORMAT
            amount_cell.alignment = Alignment(horizontal="right")
            for col in range(1, len(COLUMNS) + 1):
                ws.cell(row=row, column=col).border = _thin_border("bottom")
            subtotal += float(e["amount"])
            row += 1

        # Fila de subtotal del grupo
        ws.merge_cells(f"A{row}:E{row}")
        subtotal_label = ws.cell(row=row, column=1, value=f"Subtotal {pretty_label(type_key)}")
        subtotal_label.font = Font(bold=True)
        subtotal_label.alignment = Alignment(horizontal="right")
        subtotal_cell = ws.cell(row=row, column=6, value=subtotal)
        subtotal_cell.number_format = CURRENCY_FORMAT
        subtotal_cell.font = Font(bold=True)
        subtotal_cell.border = Border(top=Side(style="thin", color="98A2B3"))
        subtotal_label.border = Border(top=Side(style="thin", color="98A2B3"))
        grand_total += subtotal
        row += 2  # fila en blanco entre grupos

    # --- Total general ---
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    total_label = ws.cell(row=row, column=1, value="TOTAL GENERAL")
    total_label.font = Font(bold=True, size=12, color=COLOR_HEADER_TEXT)
    total_label.fill = PatternFill("solid", fgColor=COLOR_TOTAL_FILL)
    total_label.alignment = Alignment(horizontal="right", vertical="center")

    total_cell = ws.cell(row=row, column=6, value=grand_total)
    total_cell.number_format = CURRENCY_FORMAT
    total_cell.font = Font(bold=True, size=12, color=COLOR_HEADER_TEXT)
    total_cell.fill = PatternFill("solid", fgColor=COLOR_TOTAL_FILL)
    total_cell.alignment = Alignment(horizontal="right", vertical="center")

    if not any(by_type.values()):
        ws.cell(row=header_row + 1, column=1, value="No hay gastos para los filtros seleccionados.").font = Font(italic=True, color=COLOR_GRAY)

    # --- Anchos de columna ---
    for idx, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.sheet_view.showGridLines = False

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


COMMISSION_COLUMN_WIDTHS = [40, 14, 18]  # Ruta, Viajes, Comisión
COMMISSION_COLUMNS = ["Ruta", "Viajes", "Comisión"]


def build_commissions_workbook(drivers, company_name, month):
    """Construye un workbook con el reporte mensual de comisiones de
    conductores, agrupado por conductor con subtotal y total general.
    `drivers` es una lista de dicts: {driver_name, routes: [fila con
    origin, destination, trip_count, route_commission], trip_count,
    total_commission}."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Comisiones"

    last_col_letter = get_column_letter(len(COMMISSION_COLUMNS))

    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = f"{company_name} — Comisiones de conductores"
    ws["A1"].font = Font(bold=True, size=14, color=COLOR_PRIMARY)

    ws.merge_cells(f"A2:{last_col_letter}2")
    generated = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws["A2"] = f"Generado el {generated}  ·  Mes: {month}"
    ws["A2"].font = Font(italic=True, size=10, color=COLOR_GRAY)

    header_row = 4
    for idx, title in enumerate(COMMISSION_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=idx, value=title)
        cell.font = Font(bold=True, color=COLOR_HEADER_TEXT)
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="right" if title != "Ruta" else "left", vertical="center")
        cell.border = _thin_border("all")

    ws.freeze_panes = f"A{header_row + 1}"

    row = header_row + 1
    grand_total = 0.0
    grand_trips = 0

    for d in sorted(drivers, key=lambda x: x["driver_name"]):
        ws.merge_cells(f"A{row}:{last_col_letter}{row}")
        group_cell = ws.cell(row=row, column=1, value=d["driver_name"])
        group_cell.font = Font(bold=True, color=COLOR_PRIMARY)
        group_cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY_SOFT)
        row += 1

        for r in d["routes"]:
            ws.cell(row=row, column=1, value=f"{r['origin']} → {r['destination']}")
            trips_cell = ws.cell(row=row, column=2, value=int(r["trip_count"]))
            trips_cell.alignment = Alignment(horizontal="right")
            amount_cell = ws.cell(row=row, column=3, value=float(r["route_commission"] or 0))
            amount_cell.number_format = CURRENCY_FORMAT
            amount_cell.alignment = Alignment(horizontal="right")
            for col in range(1, len(COMMISSION_COLUMNS) + 1):
                ws.cell(row=row, column=col).border = _thin_border("bottom")
            row += 1

        ws.merge_cells(f"A{row}:A{row}")
        subtotal_label = ws.cell(row=row, column=1, value=f"Subtotal {d['driver_name']}")
        subtotal_label.font = Font(bold=True)
        subtotal_label.alignment = Alignment(horizontal="right")
        subtotal_trips = ws.cell(row=row, column=2, value=int(d["trip_count"]))
        subtotal_trips.font = Font(bold=True)
        subtotal_trips.alignment = Alignment(horizontal="right")
        subtotal_cell = ws.cell(row=row, column=3, value=float(d["total_commission"]))
        subtotal_cell.number_format = CURRENCY_FORMAT
        subtotal_cell.font = Font(bold=True)
        subtotal_cell.border = Border(top=Side(style="thin", color="98A2B3"))
        subtotal_label.border = Border(top=Side(style="thin", color="98A2B3"))
        subtotal_trips.border = Border(top=Side(style="thin", color="98A2B3"))
        grand_total += d["total_commission"]
        grand_trips += d["trip_count"]
        row += 2

    row += 1
    total_label = ws.cell(row=row, column=1, value="TOTAL GENERAL")
    total_label.font = Font(bold=True, size=12, color=COLOR_HEADER_TEXT)
    total_label.fill = PatternFill("solid", fgColor=COLOR_TOTAL_FILL)
    total_label.alignment = Alignment(horizontal="right", vertical="center")

    total_trips_cell = ws.cell(row=row, column=2, value=grand_trips)
    total_trips_cell.font = Font(bold=True, size=12, color=COLOR_HEADER_TEXT)
    total_trips_cell.fill = PatternFill("solid", fgColor=COLOR_TOTAL_FILL)
    total_trips_cell.alignment = Alignment(horizontal="right", vertical="center")

    total_cell = ws.cell(row=row, column=3, value=grand_total)
    total_cell.number_format = CURRENCY_FORMAT
    total_cell.font = Font(bold=True, size=12, color=COLOR_HEADER_TEXT)
    total_cell.fill = PatternFill("solid", fgColor=COLOR_TOTAL_FILL)
    total_cell.alignment = Alignment(horizontal="right", vertical="center")

    if not drivers:
        ws.cell(row=header_row + 1, column=1, value="No hay viajes con conductor asignado para el mes seleccionado.").font = Font(italic=True, color=COLOR_GRAY)

    for idx, width in enumerate(COMMISSION_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.sheet_view.showGridLines = False

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# Anchos aproximados para las 16 columnas de RESUMEN_COLUMNS (ver
# app/accounting.py): Origen, Num.Voucher, Fecha Liq., Cuenta, Monto Debe,
# Monto Haber, Moneda, T.Cambio, Doc, Num.Doc, Fec.Doc, Fec.Ven, RUC/DNI,
# Glosa, RUC/DNI, R. Social.
LIQUIDACION_COLUMN_WIDTHS = [9, 12, 16, 10, 13, 13, 10, 10, 7, 16, 12, 12, 14, 34, 14, 34]


def build_liquidacion_workbook(rows, company_name, filter_description):
    """Construye el workbook de "liquidación contable" en el formato EXACTO
    de la hoja resumen de la plantilla real de Harraso: una sola tabla
    plana con estas 16 columnas (mismos nombres, mismo orden — ver
    RESUMEN_COLUMNS en app/accounting.py), sin agrupar por conductor ni por
    tipo, tal como la usan para pegarla directo en su sistema contable.

    `rows` es una lista de dicts ya armados por
    app/routes/liquidaciones.py::_liquidacion_rows(), con una fila "Haber" por
    cada anticipo liquidado y una fila "Debe" por cada gasto documentado
    vinculado a ese anticipo. Claves esperadas por fila: origen,
    num_voucher, fecha_liquidacion, cuenta, monto_debe, monto_haber,
    moneda, tipo_cambio, doc, num_doc, fec_doc, fec_ven, ruc_dni, glosa,
    ruc_dni2, razon_social (cualquiera puede venir None/vacío)."""
    from app.accounting import RESUMEN_COLUMNS

    wb = Workbook()
    ws = wb.active
    ws.title = "hoja resumen"

    n_cols = len(RESUMEN_COLUMNS)
    last_col_letter = get_column_letter(n_cols)

    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = f"{company_name} — Liquidación contable"
    ws["A1"].font = Font(bold=True, size=14, color=COLOR_PRIMARY)

    ws.merge_cells(f"A2:{last_col_letter}2")
    generated = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws["A2"] = f"Generado el {generated}  ·  {filter_description}"
    ws["A2"].font = Font(italic=True, size=10, color=COLOR_GRAY)

    header_row = 4
    for idx, title in enumerate(RESUMEN_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=idx, value=title)
        cell.font = Font(bold=True, color=COLOR_HEADER_TEXT)
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        cell.alignment = Alignment(vertical="center")
        cell.border = _thin_border("all")

    ws.freeze_panes = f"A{header_row + 1}"

    money_cols = {5, 6}  # Monto Debe, Monto Haber
    rate_col = 8  # T.Cambio

    row = header_row + 1
    total_debe = 0.0
    total_haber = 0.0
    for r in rows:
        values = [
            r.get("origen") or "",
            r.get("num_voucher") or "",
            r.get("fecha_liquidacion") or "",
            r.get("cuenta") or "",
            r.get("monto_debe"),
            r.get("monto_haber"),
            r.get("moneda") or "",
            r.get("tipo_cambio"),
            r.get("doc") or "",
            r.get("num_doc") or "",
            r.get("fec_doc") or "",
            r.get("fec_ven") or "",
            r.get("ruc_dni") or "",
            r.get("glosa") or "",
            r.get("ruc_dni2") or "",
            r.get("razon_social") or "",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if col in money_cols and value is not None:
                cell.number_format = CURRENCY_FORMAT
            elif col == rate_col and value is not None:
                cell.number_format = "0.000"
            cell.border = _thin_border("bottom")
        total_debe += float(r.get("monto_debe") or 0)
        total_haber += float(r.get("monto_haber") or 0)
        row += 1

    row += 1
    total_label = ws.cell(row=row, column=1, value="TOTALES")
    total_label.font = Font(bold=True, size=12, color=COLOR_HEADER_TEXT)
    total_label.fill = PatternFill("solid", fgColor=COLOR_TOTAL_FILL)
    ws.merge_cells(f"A{row}:D{row}")
    total_label.alignment = Alignment(horizontal="right", vertical="center")

    total_debe_cell = ws.cell(row=row, column=5, value=total_debe)
    total_debe_cell.number_format = CURRENCY_FORMAT
    total_haber_cell = ws.cell(row=row, column=6, value=total_haber)
    total_haber_cell.number_format = CURRENCY_FORMAT
    for c in (total_debe_cell, total_haber_cell):
        c.font = Font(bold=True, size=12, color=COLOR_HEADER_TEXT)
        c.fill = PatternFill("solid", fgColor=COLOR_TOTAL_FILL)
        c.alignment = Alignment(horizontal="right", vertical="center")
    for col in range(7, n_cols + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=COLOR_TOTAL_FILL)

    if not rows:
        ws.cell(row=header_row + 1, column=1, value="No hay liquidaciones para los filtros seleccionados.").font = Font(italic=True, color=COLOR_GRAY)

    for idx, width in enumerate(LIQUIDACION_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.sheet_view.showGridLines = False

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


GPS_DAILY_COLUMNS = ["Unidad", "Horas manejadas", "Km avanzados"]
GPS_DAILY_COLUMN_WIDTHS = [16, 18, 16]


def build_gps_daily_workbook(rows, company_name, date):
    """Construye el reporte diario de GPS (31 ago, pedido de Braulio):
    horas manejadas y km avanzados por unidad en un día puntual, calculados
    a partir del historial de posiciones (ver app/gps_stats.py). `rows` es
    una lista de dicts {plate, hours, km}."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte GPS diario"

    last_col_letter = get_column_letter(len(GPS_DAILY_COLUMNS))

    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = f"{company_name} — Reporte diario de GPS"
    ws["A1"].font = Font(bold=True, size=14, color=COLOR_PRIMARY)

    ws.merge_cells(f"A2:{last_col_letter}2")
    generated = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws["A2"] = f"Generado el {generated}  ·  Día: {date}"
    ws["A2"].font = Font(italic=True, size=10, color=COLOR_GRAY)

    header_row = 4
    for idx, title in enumerate(GPS_DAILY_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=idx, value=title)
        cell.font = Font(bold=True, color=COLOR_HEADER_TEXT)
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="right" if title != "Unidad" else "left", vertical="center")
        cell.border = _thin_border("all")

    ws.freeze_panes = f"A{header_row + 1}"

    row = header_row + 1
    total_hours = 0.0
    total_km = 0.0
    for r in sorted(rows, key=lambda x: x["plate"]):
        ws.cell(row=row, column=1, value=r["plate"])
        hours_cell = ws.cell(row=row, column=2, value=round(float(r["hours"] or 0), 1))
        hours_cell.alignment = Alignment(horizontal="right")
        km_cell = ws.cell(row=row, column=3, value=round(float(r["km"] or 0), 1))
        km_cell.alignment = Alignment(horizontal="right")
        for col in range(1, len(GPS_DAILY_COLUMNS) + 1):
            ws.cell(row=row, column=col).border = _thin_border("bottom")
        total_hours += r["hours"] or 0
        total_km += r["km"] or 0
        row += 1

    if not rows:
        ws.cell(row=row, column=1, value="No hay unidades registradas.").font = Font(italic=True, color=COLOR_GRAY)
        row += 1

    row += 1
    total_label = ws.cell(row=row, column=1, value="TOTAL GENERAL")
    total_label.font = Font(bold=True, size=12, color=COLOR_HEADER_TEXT)
    total_label.fill = PatternFill("solid", fgColor=COLOR_TOTAL_FILL)
    total_label.alignment = Alignment(horizontal="right", vertical="center")

    total_hours_cell = ws.cell(row=row, column=2, value=round(total_hours, 1))
    total_km_cell = ws.cell(row=row, column=3, value=round(total_km, 1))
    for c in (total_hours_cell, total_km_cell):
        c.font = Font(bold=True, size=12, color=COLOR_HEADER_TEXT)
        c.fill = PatternFill("solid", fgColor=COLOR_TOTAL_FILL)
        c.alignment = Alignment(horizontal="right", vertical="center")

    for idx, width in enumerate(GPS_DAILY_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.sheet_view.showGridLines = False

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
