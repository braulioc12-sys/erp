"""Importación masiva desde Excel para Flota, Conductores y Rutas (30 ago,
pedido de Braulio: "poder importar de manera masiva a través de un formato
de excel"). Este módulo trae el motor genérico (generar la plantilla
.xlsx descargable y leer/validar el archivo que el usuario sube); guardar
las filas ya validadas en la base de datos (con las reglas propias de
duplicados de cada tabla) vive en cada app/routes/<módulo>.py, en una
función `_apply_..._import(rows)` que arma el mismo dict de resultado:
{"created": int, "updated": int, "skipped": [...], "errors": [...]}
(cada entrada de skipped/errors es {"row": <fila del Excel>, "message": str}).

Diseño de la plantilla (misma estructura para los 3 módulos, ver
build_import_template): fila 1 = título, fila 2 = instrucciones, fila 3 =
encabezados, fila 4 = fila de EJEMPLO, fila 5 en adelante = donde el
usuario llena sus datos. La fila de ejemplo se detecta y se descarta sola
al importar (ver _row_matches_example) — si el usuario la deja tal cual no
pasa nada raro, simplemente no se cuenta. La búsqueda de la fila de
encabezados es dinámica (_find_header_row, primeras 15 filas) en vez de
asumir que siempre es la fila 3, para tolerar que alguien borre la fila de
título/instrucciones antes de subir el archivo."""
import io
import unicodedata
from datetime import date, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Paleta acorde al resto de exportaciones (ver app/reports.py)
COLOR_PRIMARY = "1D4ED8"
COLOR_HEADER_TEXT = "FFFFFF"
COLOR_GRAY = "667085"
COLOR_EXAMPLE_FILL = "F2F4F7"
COLOR_WARNING = "B54708"
COLOR_WARNING_SOFT = "FFFAEB"

HEADER_ROW = 3
EXAMPLE_ROW = 4
DATA_START_ROW = 5
TEMPLATE_DATA_ROWS = 500  # filas con validación de lista desplegable ya lista para llenar


class ImportColumn:
    """Especifica una columna de la plantilla de importación.

    kind: "text" | "float" | "date" | "choice"
    choices: para kind="choice", lista de (código_guardado, [alias_aceptados,...]);
        el valor de la celda se compara sin distinguir mayúsculas/acentos
        contra el código y sus alias.
    """

    __slots__ = ("key", "header", "kind", "required", "choices", "note", "width")

    def __init__(self, key, header, kind="text", required=False, choices=None, note="", width=22):
        self.key = key
        self.header = header
        self.kind = kind
        self.required = required
        self.choices = choices or []
        self.note = note
        self.width = width


def _strip_accents(value):
    return "".join(c for c in unicodedata.normalize("NFD", value) if unicodedata.category(c) != "Mn")


def _normalize(value):
    """Para comparar encabezados/valores sin distinguir mayúsculas ni
    acentos. También quita un "*" final (y el espacio antes) porque el
    encabezado de una columna obligatoria en la plantilla se escribe como
    "Placa *" (ver build_import_template) y debe seguir emparejando con la
    columna "Placa" de la especificación."""
    text = _strip_accents(str(value).strip().lower())
    return text.rstrip("*").strip()


def _thin_border():
    side = Side(style="thin", color="D0D5DD")
    return Border(left=side, right=side, top=side, bottom=side)


def parse_excel_float(raw):
    if isinstance(raw, (int, float)):
        return float(raw)
    try:
        return float(str(raw).strip().replace(",", ""))
    except (TypeError, ValueError):
        return None


def parse_excel_date(raw):
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")
    if isinstance(raw, date):
        return raw.strftime("%Y-%m-%d")
    text = str(raw).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def normalize_choice(raw, choices):
    norm = _normalize(raw)
    for code, aliases in choices:
        if norm == _normalize(code):
            return code
        for alias in aliases:
            if norm == _normalize(alias):
                return code
    return None


def _convert_cell(raw, column):
    is_blank = raw is None or (isinstance(raw, str) and raw.strip() == "")
    if is_blank:
        return ("" if column.kind == "text" else None), None
    if column.kind == "text":
        return str(raw).strip(), None
    if column.kind == "float":
        value = parse_excel_float(raw)
        if value is None:
            return None, f"valor numérico inválido ('{raw}'); se dejó vacío."
        return value, None
    if column.kind == "date":
        value = parse_excel_date(raw)
        if value is None:
            return None, f"fecha inválida ('{raw}'), usa el formato AAAA-MM-DD; se dejó vacía."
        return value, None
    if column.kind == "choice":
        value = normalize_choice(raw, column.choices)
        if value is None:
            valid = ", ".join(code for code, _ in column.choices)
            return None, f"valor no reconocido ('{raw}'); valores válidos: {valid}. Se dejó vacío."
        return value, None
    return str(raw).strip(), None


def _find_header_row(ws, columns, max_scan_rows=15):
    """Busca en las primeras filas cuál trae los encabezados de columna,
    en vez de asumir que siempre es la fila HEADER_ROW: así tolera que el
    usuario borre o mueva la fila de título/instrucciones antes de
    importar. Devuelve (número_de_fila, {header_normalizado: columna_excel})
    de la mejor fila encontrada, o (None, {}) si ninguna fila trae ni una
    columna reconocible."""
    required_norms = {_normalize(c.header) for c in columns if c.required}
    all_norms = {_normalize(c.header) for c in columns}
    best_row, best_index, best_score = None, {}, -1
    scan_limit = min(ws.max_row or 1, max_scan_rows)
    for r in range(1, scan_limit + 1):
        header_index = {}
        for cell in ws[r]:
            if cell.value is None:
                continue
            norm = _normalize(str(cell.value))
            if norm in all_norms:
                header_index[norm] = cell.column
        if required_norms and required_norms.issubset(header_index.keys()):
            return r, header_index
        if len(header_index) > best_score:
            best_row, best_index, best_score = r, header_index, len(header_index)
    if best_score <= 0:
        return None, {}
    return best_row, best_index


def _row_matches_example(row_data, example_values, columns):
    """True si TODAS las columnas obligatorias de la fila coinciden
    exactamente (sin distinguir mayúsculas) con los valores de la fila de
    ejemplo de la plantilla — se usa para descartarla sola si el usuario
    la deja sin borrar."""
    if not example_values:
        return False
    for col in columns:
        if not col.required:
            continue
        example_value = example_values.get(col.key)
        if example_value in (None, ""):
            continue
        actual = row_data.get(col.key)
        if actual is None or str(actual).strip().lower() != str(example_value).strip().lower():
            return False
    return True


def read_import_rows(file_storage, columns, example_values=None):
    """Lee y valida el archivo subido. Devuelve (rows, file_error,
    example_skips): si file_error no es None, el archivo no se pudo
    procesar en absoluto (rows viene vacío) y file_error es el mensaje a
    mostrar. Si no hay error, rows es la lista de filas con datos (dicts
    con una clave por columna más "_row_number" y "_warnings"), y
    example_skips es la lista de números de fila que se reconocieron como
    la fila de ejemplo de la plantilla y se descartaron solas."""
    if not file_storage or not file_storage.filename:
        return [], "No se seleccionó ningún archivo.", []
    if not file_storage.filename.lower().endswith(".xlsx"):
        return [], (
            "El archivo debe estar en formato Excel (.xlsx). Si lo tienes en otro formato, "
            "ábrelo en Excel y usa \"Guardar como\" → .xlsx."
        ), []
    try:
        wb = load_workbook(io.BytesIO(file_storage.read()), data_only=True)
    except Exception:
        return [], "No se pudo leer el archivo. Verifica que sea un Excel válido y no esté dañado.", []

    ws = wb["Datos"] if "Datos" in wb.sheetnames else wb.worksheets[0]

    header_row, header_index = _find_header_row(ws, columns)
    if header_row is None:
        return [], (
            "No se encontró una fila de encabezados reconocible. Descarga la plantilla y no cambies "
            "los nombres de columna."
        ), []

    missing_required = [c.header for c in columns if c.required and _normalize(c.header) not in header_index]
    if missing_required:
        return [], (
            "Al archivo le faltan estas columnas obligatorias: " + ", ".join(missing_required) +
            ". Descarga la plantilla nuevamente y no cambies los nombres de columna."
        ), []

    col_idx_by_key = {}
    for col in columns:
        norm = _normalize(col.header)
        if norm in header_index:
            col_idx_by_key[col.key] = header_index[norm]

    rows = []
    example_skips = []
    for r in range(header_row + 1, ws.max_row + 1):
        row_data = {}
        warnings = []
        any_value = False
        for col in columns:
            idx = col_idx_by_key.get(col.key)
            raw = ws.cell(row=r, column=idx).value if idx else None
            if raw is not None and str(raw).strip() != "":
                any_value = True
            value, warn = _convert_cell(raw, col)
            row_data[col.key] = value
            if warn:
                warnings.append(f"Columna \"{col.header}\": {warn}")
        if not any_value:
            continue
        if _row_matches_example(row_data, example_values, columns):
            example_skips.append(r)
            continue
        row_data["_row_number"] = r
        row_data["_warnings"] = warnings
        rows.append(row_data)

    return rows, None, example_skips


def build_import_template(title, columns, example_values):
    """Genera el archivo .xlsx descargable: hoja "Datos" (título +
    instrucciones + encabezados + fila de ejemplo + validación de listas
    desplegables para las columnas de tipo choice) y hoja "Instrucciones"
    con el detalle de cada columna."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    n_cols = len(columns)
    last_col = get_column_letter(n_cols)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"Plantilla de importación — {title}"
    ws["A1"].font = Font(bold=True, size=14, color=COLOR_PRIMARY)

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = (
        f"Completa los datos desde la fila {DATA_START_ROW} (una fila por registro). "
        f"La fila {EXAMPLE_ROW} es un EJEMPLO: bórrala o sobrescríbela con datos reales — si la dejas tal "
        f"cual, se descarta sola al importar. No cambies los nombres de columna de la fila {HEADER_ROW}. "
        "Revisa la hoja \"Instrucciones\" para el detalle de cada columna."
    )
    ws["A2"].font = Font(italic=True, size=10, color=COLOR_GRAY)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 32

    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=HEADER_ROW, column=idx, value=col.header + (" *" if col.required else ""))
        cell.font = Font(bold=True, color=COLOR_HEADER_TEXT)
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = _thin_border()

    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=EXAMPLE_ROW, column=idx, value=example_values.get(col.key, ""))
        cell.font = Font(italic=True, color=COLOR_GRAY)
        cell.fill = PatternFill("solid", fgColor=COLOR_EXAMPLE_FILL)
        cell.border = _thin_border()

    ws.freeze_panes = f"A{DATA_START_ROW}"

    for idx, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = col.width
        if col.kind == "choice" and col.choices:
            valid_values = [code for code, _ in col.choices]
            dv = DataValidation(
                type="list",
                formula1='"' + ",".join(valid_values) + '"',
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Valor no válido",
                error="Usa uno de: " + ", ".join(valid_values),
            )
            col_letter = get_column_letter(idx)
            dv.add(f"{col_letter}{DATA_START_ROW}:{col_letter}{DATA_START_ROW + TEMPLATE_DATA_ROWS}")
            ws.add_data_validation(dv)

    ws.sheet_view.showGridLines = False

    # --- Hoja de instrucciones ---
    ws2 = wb.create_sheet("Instrucciones")
    ws2.merge_cells("A1:C1")
    ws2["A1"] = f"Instrucciones — {title}"
    ws2["A1"].font = Font(bold=True, size=13, color=COLOR_PRIMARY)

    header2 = ["Columna", "Obligatoria", "Formato / valores válidos"]
    for idx, h in enumerate(header2, start=1):
        cell = ws2.cell(row=3, column=idx, value=h)
        cell.font = Font(bold=True, color=COLOR_HEADER_TEXT)
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
        cell.border = _thin_border()

    row = 4
    for col in columns:
        ws2.cell(row=row, column=1, value=col.header).border = _thin_border()
        ws2.cell(row=row, column=2, value="Sí" if col.required else "No").border = _thin_border()
        note = col.note or ""
        if col.kind == "choice" and col.choices:
            valid = ", ".join(code for code, _ in col.choices)
            note = (note + " " if note else "") + f"Valores válidos: {valid}."
        if col.kind == "date":
            note = (note + " " if note else "") + "Formato de fecha: AAAA-MM-DD (ej. 2026-12-31)."
        cell = ws2.cell(row=row, column=3, value=note)
        cell.border = _thin_border()
        for c in range(1, 4):
            ws2.cell(row=row, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 74
    ws2.sheet_view.showGridLines = False

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# --- Especificaciones de columnas por módulo ---

VEHICLE_COLUMNS = [
    ImportColumn("plate", "Placa", kind="text", required=True, width=14,
                 note="Debe ser única. Si la placa ya existe en Flota no se crea de nuevo, pero si la fila "
                      "trae 'ID en el proveedor de GPS' se actualiza ese dato en la unidad existente (el "
                      "resto de los campos de esa fila no se toca)."),
    ImportColumn("brand", "Marca", kind="text", width=16),
    ImportColumn("model", "Modelo", kind="text", width=16),
    ImportColumn(
        "vehicle_type", "Tipo de unidad", kind="choice", width=16,
        choices=[
            ("CAMION", ["CAMION", "CAMIÓN", "CAMION SIMPLE"]),
            ("TRACTO", ["TRACTO", "TRACTOCAMION", "TRACTO CAMION", "CABEZAL"]),
            ("CARRETA", ["CARRETA", "SEMIRREMOLQUE"]),
        ],
        note="Si se deja vacío, se usa CAMION. Determina el diagrama de neumáticos de la unidad.",
    ),
    ImportColumn(
        "status", "Estado", kind="choice", width=16,
        choices=[
            ("ACTIVO", ["ACTIVO", "ACTIVA"]),
            ("MANTENIMIENTO", ["MANTENIMIENTO", "EN MANTENIMIENTO"]),
            ("INACTIVO", ["INACTIVO", "INACTIVA"]),
        ],
        note="Si se deja vacío, se usa ACTIVO.",
    ),
    ImportColumn("owner", "Propietario", kind="text", width=20,
                 note="Texto libre; usa el mismo nombre que en Catálogos → Propietarios de unidades si quieres que coincida."),
    ImportColumn("capacity_kg", "Capacidad (kg)", kind="float", width=14),
    ImportColumn("current_km", "Kilometraje actual", kind="float", width=16),
    ImportColumn("soat_expiry", "Vencimiento SOAT", kind="date", width=18),
    ImportColumn("technical_review_expiry", "Vencimiento revisión técnica", kind="date", width=22),
    ImportColumn("notes", "Notas", kind="text", width=30),
    ImportColumn(
        "gps_external_id", "ID en el proveedor de GPS", kind="text", width=26,
        note="Opcional. Tal como aparece en Frotcom (Configuración → Integraciones). Si la placa ya "
             "existe en Flota, este es el único campo de la fila que se actualiza; si la placa es "
             "nueva, la unidad se crea con este ID puesto desde el inicio.",
    ),
]

VEHICLE_EXAMPLE = {
    "plate": "EJM-000",
    "brand": "Volvo",
    "model": "FH 480",
    "vehicle_type": "CAMION",
    "status": "ACTIVO",
    "owner": "",
    "capacity_kg": 15000,
    "current_km": 120000,
    "soat_expiry": "2026-12-31",
    "technical_review_expiry": "2026-12-31",
    "notes": "Fila de ejemplo — bórrala o sobrescríbela",
    "gps_external_id": "190119",
}

DRIVER_COLUMNS = [
    ImportColumn("name", "Nombre completo", kind="text", required=True, width=26),
    ImportColumn("document_number", "DNI", kind="text", width=14,
                 note="Si coincide con el DNI de un conductor ya registrado, esa fila se omite (no se sobrescribe)."),
    ImportColumn("license_number", "N° de licencia", kind="text", width=16),
    ImportColumn("license_expiry", "Vencimiento de brevete", kind="date", width=18),
    ImportColumn("medical_exam_date", "Examen médico — fecha", kind="date", width=18),
    ImportColumn("medical_exam_expiry", "Examen médico — vencimiento", kind="date", width=20),
    ImportColumn("backus_driving_exam_date", "Examen de manejo Backus — fecha", kind="date", width=22),
    ImportColumn("backus_driving_exam_expiry", "Examen de manejo Backus — vencimiento", kind="date", width=24),
    ImportColumn("backus_training_date", "Capacitación Plan de tráfico Backus — fecha", kind="date", width=26),
    ImportColumn("backus_training_expiry", "Capacitación Plan de tráfico Backus — vencimiento", kind="date", width=28),
    ImportColumn("dds_date", "Escuela de conductores — fecha", kind="date", width=22),
    ImportColumn("dds_expiry", "Escuela de conductores — vencimiento", kind="date", width=24),
    ImportColumn("phone", "Teléfono", kind="text", width=14),
    ImportColumn(
        "status", "Estado", kind="choice", width=14,
        choices=[("ACTIVO", ["ACTIVO", "ACTIVA"]), ("INACTIVO", ["INACTIVO", "INACTIVA"])],
        note="Si se deja vacío, se usa ACTIVO.",
    ),
]

DRIVER_EXAMPLE = {
    "name": "Nombre de Ejemplo",
    "document_number": "00000000",
    "license_number": "Q00000000",
    "license_expiry": "2027-01-01",
    "status": "ACTIVO",
}

ROUTE_COLUMNS = [
    ImportColumn("origin", "Origen", kind="text", required=True, width=20),
    ImportColumn("destination", "Destino", kind="text", required=True, width=20,
                 note="La combinación Origen + Destino debe ser única; si ya existe, se actualizan sus montos."),
    ImportColumn("default_expense_amount", "Viáticos (S/)", kind="float", width=16),
    ImportColumn("default_commission_amount", "Comisión conductor (S/)", kind="float", width=22),
]

ROUTE_EXAMPLE = {
    "origin": "Lima",
    "destination": "Trujillo (ejemplo)",
    "default_expense_amount": 350,
    "default_commission_amount": 80,
}

# 3 sep, pedido de Braulio: carga masiva del último cambio de aceite por
# placa, a partir de un Excel de terceros (ver app/routes/flota.py
# import_oil_changes()). A diferencia de VEHICLE_COLUMNS, esta importación
# es SOLO para actualizar unidades que YA existen en Flota — decisión
# explícita de Braulio ("no crear nada, solo actualizar las que ya
# existen"), así que no incluye los campos que se necesitarían para crear
# una unidad nueva (marca, tipo, capacidad, etc.); ver
# _apply_oil_change_import() para el detalle de placas no encontradas.
OIL_CHANGE_COLUMNS = [
    ImportColumn("plate", "Placa", kind="text", required=True, width=14,
                 note="Debe ser una placa YA registrada en Flota. Si no existe, la fila se omite (esta "
                      "importación no crea unidades nuevas)."),
    ImportColumn("oil_change_km", "Kilometraje último cambio", kind="float", width=20),
    ImportColumn("oil_change_date", "Fecha último cambio", kind="date", width=18),
    ImportColumn("workshop", "Taller donde se realizó", kind="text", width=24),
    ImportColumn("oil_type", "Aceite", kind="text", width=18),
]

OIL_CHANGE_EXAMPLE = {
    "plate": "EJM-000",
    "oil_change_km": 120000,
    "oil_change_date": "2026-07-01",
    "workshop": "Taller de ejemplo",
    "oil_type": "15W40",
}
